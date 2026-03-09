"""
push_to_github.py
=================
Exports shipments from the Excel database to shipments.json,
then pushes it to your GitHub repo so the Streamlit web app updates.

Called automatically by process_dhl_pdfs.py after each run.

Setup (one time):
    pip install PyGithub openpyxl
    Set environment variable:  GITHUB_TOKEN=your_personal_access_token
    Edit GITHUB_REPO below to match your repo name.
"""

import json
import os
import base64
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook

# ── CONFIG — edit these ───────────────────────────────────────────────────────
GITHUB_REPO  = "Gofetch-11/dhl-mahoney-dashboard"  # your GitHub repo
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")       # set as env variable
XLSX_PATH    = r"C:\Users\MatthewAMahoney\Dropbox\DHLSD - NEW Job Tickets\DHL_Mahoney_Shipment_Database.xlsx"
# ─────────────────────────────────────────────────────────────────────────────


def xlsx_to_json(xlsx_path: str) -> list[dict]:
    """Read Master Database sheet and return list of row dicts."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Master Database"]
    headers = [cell.value for cell in ws[3]]
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]:
            continue
        d = {}
        for h, v in zip(headers, row):
            if h and h != "Source PDF":
                d[str(h)] = str(v) if v is not None else ""
        rows.append(d)
    return rows


def push_to_github(rows: list[dict]) -> bool:
    """Push shipments.json to GitHub repo via API."""
    try:
        from github import Github, GithubException
    except ImportError:
        print("  ⚠ PyGithub not installed. Run: pip install PyGithub")
        return False

    if not GITHUB_TOKEN:
        print("  ⚠ GITHUB_TOKEN not set. Export it as an environment variable.")
        print("     Windows: setx GITHUB_TOKEN your_token_here")
        return False

    try:
        g = Github(GITHUB_TOKEN)
        repo = g.get_repo(GITHUB_REPO)

        content = json.dumps(rows, indent=2, ensure_ascii=False)
        encoded = content.encode("utf-8")
        commit_msg = f"Update shipments.json — {len(rows)} jobs — {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        # Update or create the file
        try:
            existing = repo.get_contents("shipments.json")
            repo.update_file("shipments.json", commit_msg, content, existing.sha)
            print(f"  ✓ GitHub updated — {len(rows)} shipments pushed")
        except GithubException:
            repo.create_file("shipments.json", commit_msg, content)
            print(f"  ✓ GitHub created shipments.json — {len(rows)} shipments")

        return True

    except Exception as e:
        print(f"  ✗ GitHub push failed: {e}")
        return False


def run(xlsx_path: str = XLSX_PATH):
    print(f"  Reading {xlsx_path}…")
    rows = xlsx_to_json(xlsx_path)
    print(f"  Found {len(rows)} shipments")
    push_to_github(rows)


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", default=XLSX_PATH)
    args = p.parse_args()
    run(args.xlsx)
